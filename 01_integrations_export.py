"""
01_integrations_export.py
=========================
Extracts ALL OIC Gen3 integration details → Excel workbook.

Sheets produced:
  1. Integrations  — master list with key fields
  2. Summary       — counts by status, style, pattern

API used:
  GET /ic/api/integration/v1/integrations
  GET /ic/api/integration/v1/integrations/{id}  (for dependency details)
"""

import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from oic_client import OICClient

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

OUTPUT_FILE = "OIC_Integrations_Report.xlsx"


def main():
    client = OICClient()

    # ── Fetch all integrations (list endpoint returns most fields) ──
    logger.info("Fetching all integrations...")
    integrations = client.get_all_integrations(
        limit=100,
        expand="connection",  # include connection details in list response
        order_by="name",
    )
    logger.info("Retrieved %d integrations", len(integrations))

    # ── Build master DataFrame ──
    rows = []
    for intg in integrations:
        # The list response uses camelCase in the actual JSON payload
        # (despite the schema doc showing hyphenated names)
        intg_id = intg.get("id", "")
        code = intg.get("code", intg_id.split("|")[0] if "|" in intg_id else intg_id)
        version = intg.get("version", intg_id.split("|")[1] if "|" in intg_id else "")

        # Connections used (from expand=connection)
        endpoints = intg.get("endPoints", intg.get("end-points", []))
        conn_list = []
        for ep in (endpoints or []):
            conn = ep.get("connection", {})
            conn_id = conn.get("id", "")
            conn_name = ep.get("name", "")
            conn_role = ep.get("role", "")
            if conn_id:
                conn_list.append(f"{conn_id} ({conn_role})")

        # Dependencies (if present in list response)
        deps = intg.get("dependencies", {})
        dep_lookups = [lk.get("name", "") for lk in (deps.get("lookups", []) or [])]
        dep_connections = [cn.get("id", "") for cn in (deps.get("connections", []) or [])]
        dep_libraries = [lb.get("code", lb.get("display-name", ""))
                         for lb in (deps.get("libraries", []) or [])]

        rows.append({
            "Code": code,
            "Version": version,
            "Name": intg.get("name", ""),
            "Description": intg.get("description", ""),
            "Status": intg.get("status", intg.get("activation-status", "")),
            "Pattern": intg.get("pattern", intg.get("patternDescription", "")),
            "Style": intg.get("style", ""),
            "Style Description": intg.get("styleDescription", intg.get("style-description", "")),
            "Package": intg.get("packageName", intg.get("package-name", "")),
            "Schedule Applicable": intg.get("scheduleApplicableFlag",
                                            intg.get("schedule-applicable-flag", "")),
            "Schedule Defined": intg.get("scheduleDefinedFlag",
                                         intg.get("schedule-defined-flag", "")),
            "Tracing Enabled": intg.get("tracingEnabledFlag",
                                         intg.get("tracing-enabled-flag", "")),
            "Payload Tracing": intg.get("payloadTracingEnabledFlag",
                                         intg.get("payload-tracing-enabled-flag", "")),
            "% Complete": intg.get("percentageComplete",
                                    intg.get("percentage-complete", "")),
            "Created By": intg.get("createdBy", intg.get("created-by", "")),
            "Created": intg.get("created", ""),
            "Last Updated By": intg.get("lastUpdatedBy", intg.get("last-updated-by", "")),
            "Last Updated": intg.get("lastUpdated", intg.get("last-updated", "")),
            "Locked": intg.get("lockedFlag", intg.get("locked-flag", False)),
            "Locked By": intg.get("lockedBy", intg.get("locked-by", "")),
            "Endpoint URI": intg.get("endPointURI", intg.get("end-point-uri", "")),
            "Connections Used": " | ".join(conn_list) if conn_list else "",
            "Dep Lookups": " | ".join(dep_lookups) if dep_lookups else "",
            "Dep Connections": " | ".join(dep_connections) if dep_connections else "",
            "Dep Libraries": " | ".join(dep_libraries) if dep_libraries else "",
            "Project Type": intg.get("projectType", intg.get("project-type", "")),
            "Compatible": intg.get("compatible", ""),
        })

    df = pd.DataFrame(rows)

    # ── Summary DataFrame ──
    summary_rows = []
    if not df.empty:
        for col in ["Status", "Style", "Pattern", "Project Type"]:
            counts = df[col].value_counts()
            for val, cnt in counts.items():
                summary_rows.append({"Category": col, "Value": val, "Count": cnt})
        summary_rows.append({"Category": "TOTAL", "Value": "All Integrations", "Count": len(df)})
    df_summary = pd.DataFrame(summary_rows)

    # ── Write Excel ──
    wb = Workbook()

    # --- Sheet 1: Integrations ---
    ws1 = wb.active
    ws1.title = "Integrations"
    _write_df_to_sheet(ws1, df)
    _apply_status_colors(ws1, df, col_name="Status")

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")
    _write_df_to_sheet(ws2, df_summary)

    wb.save(OUTPUT_FILE)
    logger.info("Report saved: %s", OUTPUT_FILE)


# ── Excel formatting helpers ──

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9E2EC"),
)

STATUS_COLORS = {
    "ACTIVATED":        PatternFill("solid", fgColor="C6EFCE"),
    "CONFIGURED":       PatternFill("solid", fgColor="FFEB9C"),
    "INPROGRESS":       PatternFill("solid", fgColor="BDD7EE"),
    "FAILEDACTIVATION": PatternFill("solid", fgColor="FFC7CE"),
}


def _write_df_to_sheet(ws, df):
    if df.empty:
        ws.append(["No data"])
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for c_idx, _ in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _apply_status_colors(ws, df, col_name="Status"):
    if df.empty or col_name not in df.columns:
        return
    col_idx = list(df.columns).index(col_name) + 1
    for r_idx in range(2, ws.max_row + 1):
        val = str(ws.cell(row=r_idx, column=col_idx).value or "").upper()
        fill = STATUS_COLORS.get(val)
        if fill:
            ws.cell(row=r_idx, column=col_idx).fill = fill


if __name__ == "__main__":
    main()
