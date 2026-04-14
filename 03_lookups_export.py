"""
03_lookups_export.py
====================
Extracts ALL OIC Gen3 lookup details → Excel workbook.
Maps each lookup to the integrations that use it.

Strategy for lookup-to-integration mapping:
  OIC provides no direct API to query "which integrations use lookup X".
  Two-pass approach:
    Pass 1 — Check the 'dependencies.lookups[]' array in each integration detail JSON.
    Pass 2 — Download each integration's IAR archive, scan XSLT/XML files for
             dvm:lookup('LOOKUP_NAME',...) references (catches lookups embedded in mappers).

Sheets produced:
  1. Lookups         — master list with columns, row count, status
  2. Lookup Data     — actual lookup row data (columns + values)
  3. Lookup Usage    — which integrations reference each lookup
  4. Summary         — counts by status, usage

APIs used:
  GET /ic/api/integration/v1/lookups              (list all)
  GET /ic/api/integration/v1/lookups/{name}        (detail with rows/columns)
  GET /ic/api/integration/v1/integrations          (for dependency scan)
  GET /ic/api/integration/v1/integrations/{id}/archive  (IAR download for deep scan)
"""

import io
import json
import logging
import os
import re
import zipfile
import urllib.parse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from oic_client import OICClient

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

EXTS_DIR = os.path.join(os.path.dirname(__file__), "exports")
OUTPUT_FILE = os.path.join(EXTS_DIR, "OIC_Lookups_Report.xlsx")

# Set to True to enable deep IAR scanning (slower but catches XSLT-embedded lookups)
ENABLE_IAR_SCAN = False


def main():
    client = OICClient()

    # ── Create output directory ──
    os.makedirs(EXTS_DIR, exist_ok=True)

    # ── Step 1: Fetch all lookups ──
    logger.info("Fetching all lookups...")
    lookups = client.get_all_lookups(limit=100)
    logger.info("Retrieved %d lookups", len(lookups))

    # ── Step 2: Fetch detail for each lookup (to get columns + rows) ──
    lookup_details = {}
    for lk in lookups:
        lk_name = lk.get("name", "")
        if not lk_name:
            continue
        try:
            detail = client.get_lookup_detail(lk_name)
            lookup_details[lk_name] = detail
        except Exception as e:
            logger.warning("Failed to fetch detail for lookup '%s': %s", lk_name, e)
            lookup_details[lk_name] = lk  # fallback to list-level data

    # ── Step 3: Fetch integrations for usage mapping ──
    logger.info("Fetching all integrations for lookup-usage mapping...")
    integrations = client.get_all_integrations(limit=100)
    logger.info("Retrieved %d integrations", len(integrations))

    # ── Pass 1: dependency-based mapping ──
    # Check 'dependencies.lookups[]' in each integration
    usage_map = {}  # lookup_name → list of integration refs
    for intg in integrations:
        intg_id = intg.get("id", "")
        intg_name = intg.get("name", "")
        intg_status = intg.get("status", intg.get("activation-status", ""))
        intg_code = intg.get("code", intg_id.split("|")[0] if "|" in intg_id else intg_id)
        intg_version = intg.get("version", intg_id.split("|")[1] if "|" in intg_id else "")

        deps = intg.get("dependencies", {})
        dep_lookups = deps.get("lookups", []) or []
        for dep_lk in dep_lookups:
            lk_name = dep_lk.get("name", "")
            if not lk_name:
                continue
            if lk_name not in usage_map:
                usage_map[lk_name] = []
            usage_map[lk_name].append({
                "Integration Code": intg_code,
                "Integration Version": intg_version,
                "Integration Name": intg_name,
                "Integration Status": intg_status,
                "Detection Method": "API Dependency",
            })

    # ── Pass 2 (optional): IAR deep scan ──
    if ENABLE_IAR_SCAN:
        logger.info("Starting IAR deep scan for XSLT-embedded lookup references...")
        all_lookup_names = set(lookup_details.keys())
        # Build regex pattern: dvm:lookup('LOOKUP_NAME'
        dvm_pattern = re.compile(
            r"dvm:lookup\s*\(\s*['\"](" + "|".join(re.escape(n) for n in all_lookup_names) + r")['\"]",
            re.IGNORECASE,
        )

        for intg in integrations:
            intg_id = intg.get("id", "")
            intg_name = intg.get("name", "")
            intg_status = intg.get("status", intg.get("activation-status", ""))
            intg_code = intg.get("code", intg_id.split("|")[0] if "|" in intg_id else intg_id)
            intg_version = intg.get("version", intg_id.split("|")[1] if "|" in intg_id else "")

            # Skip if no lookup names to search for
            if not all_lookup_names:
                break

            encoded_id = OICClient.encode_integration_id(intg_code, intg_version)
            try:
                # Download IAR to memory
                resp = client._get(f"/integrations/{encoded_id}/archive", stream=True)
                iar_bytes = io.BytesIO(resp.content)

                with zipfile.ZipFile(iar_bytes) as zf:
                    for entry in zf.namelist():
                        # Scan .xsl, .xslt, .xml files inside the IAR
                        if entry.endswith((".xsl", ".xslt", ".xml", ".properties")):
                            try:
                                content = zf.read(entry).decode("utf-8", errors="ignore")
                                matches = dvm_pattern.findall(content)
                                for lk_name in set(matches):
                                    # Check if already found via Pass 1
                                    existing = usage_map.get(lk_name, [])
                                    already = any(
                                        u["Integration Code"] == intg_code
                                        and u["Integration Version"] == intg_version
                                        for u in existing
                                    )
                                    if not already:
                                        if lk_name not in usage_map:
                                            usage_map[lk_name] = []
                                        usage_map[lk_name].append({
                                            "Integration Code": intg_code,
                                            "Integration Version": intg_version,
                                            "Integration Name": intg_name,
                                            "Integration Status": intg_status,
                                            "Detection Method": f"IAR Scan ({entry})",
                                        })
                            except Exception:
                                pass
            except Exception as e:
                logger.warning("Failed to scan IAR for %s: %s", intg_id, e)

        logger.info("IAR deep scan complete.")

    # ── Build Lookups DataFrame ──
    lk_rows = []
    for lk_name, detail in lookup_details.items():
        columns = detail.get("columns", [])
        rows_data = detail.get("rows", detail.get("rows-as-array", []))
        row_count = detail.get("rowCount", detail.get("row-count", len(rows_data) if rows_data else 0))
        usage_count = len(usage_map.get(lk_name, []))
        active_usage = len([u for u in usage_map.get(lk_name, [])
                            if u["Integration Status"] == "ACTIVATED"])

        lk_rows.append({
            "Lookup Name": lk_name,
            "Description": detail.get("description", ""),
            "Status": detail.get("status", ""),
            "Columns": " | ".join(columns) if columns else "",
            "Column Count": len(columns),
            "Row Count": row_count,
            "Total Usage": usage_count,
            "Active Usage": active_usage,
            "Created By": detail.get("createdBy", detail.get("created-by", "")),
            "Created": detail.get("created", ""),
            "Last Updated By": detail.get("lastUpdatedBy", detail.get("last-updated-by", "")),
            "Last Updated": detail.get("lastUpdated", detail.get("last-updated", "")),
            "Locked": detail.get("lockedFlag", detail.get("locked-flag", False)),
            "Locked By": detail.get("lockedBy", detail.get("locked-by", "")),
        })
    df_lookups = pd.DataFrame(lk_rows)

    # ── Build Lookup Data DataFrame (actual values) ──
    data_rows = []
    for lk_name, detail in lookup_details.items():
        columns = detail.get("columns", [])
        rows_as_array = detail.get("rows-as-array", detail.get("rowsAsArray", []))
        rows_obj = detail.get("rows", [])

        if rows_as_array:
            # rows-as-array: each row is an array of strings matching columns order
            for row_arr in rows_as_array:
                row_dict = {"Lookup Name": lk_name}
                for idx, col_name in enumerate(columns):
                    row_dict[col_name] = row_arr[idx] if idx < len(row_arr) else ""
                data_rows.append(row_dict)
        elif rows_obj:
            # rows: each row is a SuiteLookupRowData object with fields[]
            for row_obj in rows_obj:
                row_dict = {"Lookup Name": lk_name}
                fields = row_obj.get("fields", [])
                for f in fields:
                    domain = f.get("domainName", f.get("domain-name", ""))
                    value = f.get("value", "")
                    row_dict[domain] = value
                data_rows.append(row_dict)

    df_data = pd.DataFrame(data_rows) if data_rows else pd.DataFrame(columns=["Lookup Name", "No Data"])

    # ── Build Usage DataFrame ──
    usage_rows = []
    for lk_name, usages in usage_map.items():
        for u in usages:
            usage_rows.append({"Lookup Name": lk_name, **u})
    df_usage = pd.DataFrame(usage_rows)

    # ── Build Summary DataFrame ──
    summary_rows = []
    if not df_lookups.empty:
        for col in ["Status"]:
            counts = df_lookups[col].value_counts()
            for val, cnt in counts.items():
                if val:
                    summary_rows.append({"Category": col, "Value": val, "Count": cnt})
        unused = df_lookups[df_lookups["Total Usage"] == 0]
        summary_rows.append({
            "Category": "Usage", "Value": "Unused Lookups (0 integrations)", "Count": len(unused),
        })
        summary_rows.append({
            "Category": "TOTAL", "Value": "All Lookups", "Count": len(df_lookups),
        })
        summary_rows.append({
            "Category": "Data", "Value": "Total Lookup Rows (all lookups)", "Count": len(df_data),
        })
    df_summary = pd.DataFrame(summary_rows)

    # ── Write Excel ──
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Lookups"
    _write_df_to_sheet(ws1, df_lookups)

    ws2 = wb.create_sheet("Lookup Data")
    _write_df_to_sheet(ws2, df_data)

    ws3 = wb.create_sheet("Lookup Usage")
    _write_df_to_sheet(ws3, df_usage)

    ws4 = wb.create_sheet("Summary")
    _write_df_to_sheet(ws4, df_summary)

    wb.save(OUTPUT_FILE)
    logger.info("Report saved: %s", OUTPUT_FILE)


# ── Excel formatting helpers ──

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2EC"))


def _write_df_to_sheet(ws, df):
    if df.empty:
        ws.append(["No data"])
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for c_idx, _ in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


if __name__ == "__main__":
    main()
