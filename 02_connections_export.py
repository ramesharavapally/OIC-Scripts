"""
02_connections_export.py
========================
Extracts ALL OIC Gen3 connection details → Excel workbook.
Maps each connection to the integrations that use it.

Sheets produced:
  1. Connections       — master list with properties and status
  2. Connection Usage  — which integrations use each connection
  3. Summary           — counts by adapter type, status, role

APIs used:
  GET /ic/api/integration/v1/connections           (list all)
  GET /ic/api/integration/v1/connections/{id}       (detail per connection)
  GET /ic/api/integration/v1/integrations           (with expand=connection for usage mapping)
"""

import logging
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from oic_client import OICClient

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

EXTS_DIR = os.path.join(os.path.dirname(__file__), "exports")
OUTPUT_FILE = os.path.join(EXTS_DIR, "OIC_Connections_Report.xlsx")


def main():
    client = OICClient()

    # ── Create output directory ──
    os.makedirs(EXTS_DIR, exist_ok=True)

    # ── Step 1: Fetch all connections ──
    logger.info("Fetching all connections...")
    connections = client.get_all_connections(limit=100)
    logger.info("Retrieved %d connections", len(connections))

    # ── Step 2: Fetch all integrations for usage mapping ──
    # endPoints is always present in the list response — no expand needed.
    logger.info("Fetching all integrations for connection-usage mapping...")
    integrations = client.get_all_integrations(limit=100)
    logger.info("Retrieved %d integrations", len(integrations))

    # ── Build connection → integration usage map ──
    # Key: conn_id → dict of intg_id → row  (deduped per unique integration)
    # A single integration can reference the same connection many times (multiple
    # invoke activities). We keep one entry per (conn_id, intg_id) pair so the
    # Connection Usage sheet shows one row per integration, not per activity.
    usage_map = {}   # conn_id → { intg_id: {row dict} }
    for intg in integrations:
        intg_id      = intg.get("id", "")
        intg_name    = intg.get("name", "")
        intg_status  = intg.get("status", intg.get("activation-status", ""))
        intg_code    = intg.get("code", intg_id.split("|")[0] if "|" in intg_id else intg_id)
        intg_version = intg.get("version", intg_id.split("|")[1] if "|" in intg_id else "")

        endpoints = intg.get("endPoints", intg.get("end-points", []))
        for ep in (endpoints or []):
            conn    = ep.get("connection", {})
            conn_id = conn.get("id", "")
            if not conn_id:
                continue
            if conn_id not in usage_map:
                usage_map[conn_id] = {}
            # Only add once per integration; first occurrence wins
            if intg_id not in usage_map[conn_id]:
                usage_map[conn_id][intg_id] = {
                    "Integration Code":    intg_code,
                    "Integration Version": intg_version,
                    "Integration Name":    intg_name,
                    "Integration Status":  intg_status,
                    "Role":                ep.get("role", ""),
                }

    logger.info("Usage map: %d connection(s) used across integrations", len(usage_map))

    # ── Build Connections DataFrame ──
    conn_rows = []
    for conn in connections:
        conn_id = conn.get("id", "")

        # Adapter type info (may be nested object or flat fields)
        adapter_type = conn.get("adapterType", conn.get("adapter-type", {}))
        if isinstance(adapter_type, dict):
            adapter_name = adapter_type.get("name", adapter_type.get("displayName", ""))
        else:
            adapter_name = str(adapter_type)

        # Connection properties — flatten key props
        conn_props = conn.get("connectionProperties", conn.get("connection-properties", []))
        prop_str = ""
        if conn_props:
            prop_pairs = []
            for p in conn_props:
                p_name = p.get("displayName", p.get("name", ""))
                p_val = p.get("propertyValue", p.get("value", ""))
                if p_val:
                    prop_pairs.append(f"{p_name}={p_val}")
            prop_str = " | ".join(prop_pairs)

        # Extract connection URL based on adapter type
        _ADAPTER_URL_PROP = {
            "soap": "targetWSDLURL",
            "erp":  "Host",
            "rest": "connectionUrl",
        }
        adapter_key = adapter_type.get("name", "").lower() if isinstance(adapter_type, dict) else str(adapter_type).lower()
        url_prop_name = _ADAPTER_URL_PROP.get(adapter_key, "")
        conn_url = ""
        for p in (conn_props or []):
            if p.get("propertyName", "") == url_prop_name:
                conn_url = p.get("propertyValue", p.get("value", ""))
                break

        usages = list(usage_map.get(conn_id, {}).values())
        usage_count = len(usages)
        active_usage = len([u for u in usages if u["Integration Status"] == "ACTIVATED"])
        intg_names = " | ".join(
            f"{u['Integration Name']} ({u['Integration Code']}|{u['Integration Version']}) [{u['Integration Status']}]"
            for u in usages
        )

        conn_rows.append({
            "Connection ID": conn_id,
            "Name": conn.get("name", ""),
            "Description": conn.get("description", ""),
            "Adapter Type": adapter_name,
            "Connection URL": conn_url,
            "Status": conn.get("status", ""),
            "Test Status": conn.get("testStatus", conn.get("test-status", "")),
            "Role": conn.get("role", ""),
            "Security Policy": conn.get("securityPolicy", conn.get("security-policy", "")),
            "Agent Required": conn.get("agentRequired", conn.get("agent-required", "")),
            "Agent Group": conn.get("agentGroupId", conn.get("agent-group-id", "")),
            "% Complete": conn.get("percentageComplete", conn.get("percentage-complete", "")),
            "Total Usage": usage_count,
            "Active Usage": active_usage,
            "Integrations Used In": intg_names,
            "Created By": conn.get("createdBy", conn.get("created-by", "")),
            "Created": conn.get("created", ""),
            "Last Updated By": conn.get("lastUpdatedBy", conn.get("last-updated-by", "")),
            "Last Updated": conn.get("lastUpdated", conn.get("last-updated", "")),
            "Locked": conn.get("lockedFlag", conn.get("locked-flag", False)),
            "Locked By": conn.get("lockedBy", conn.get("locked-by", "")),
            "Connection Properties": prop_str,
        })

    df_conn = pd.DataFrame(conn_rows)

    # ── Build Usage DataFrame ──
    usage_rows = []
    for conn_id, intg_dict in usage_map.items():
        conn_name = ""
        for c in conn_rows:
            if c["Connection ID"] == conn_id:
                conn_name = c["Name"]
                break
        for u in intg_dict.values():
            usage_rows.append({
                "Connection ID":   conn_id,
                "Connection Name": conn_name,
                **u,
            })
    df_usage = pd.DataFrame(usage_rows)

    # ── Build Summary DataFrame ──
    summary_rows = []
    if not df_conn.empty:
        for col in ["Adapter Type", "Status", "Role", "Security Policy"]:
            if col in df_conn.columns:
                counts = df_conn[col].value_counts()
                for val, cnt in counts.items():
                    if val:
                        summary_rows.append({"Category": col, "Value": val, "Count": cnt})

        # Unused connections
        unused = df_conn[df_conn["Total Usage"] == 0]
        summary_rows.append({
            "Category": "Usage",
            "Value": "Unused Connections (0 integrations)",
            "Count": len(unused),
        })
        summary_rows.append({
            "Category": "TOTAL",
            "Value": "All Connections",
            "Count": len(df_conn),
        })
    df_summary = pd.DataFrame(summary_rows)

    # ── Write Excel ──
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Connections"
    _write_df_to_sheet(ws1, df_conn)

    ws2 = wb.create_sheet("Connection Usage")
    _write_df_to_sheet(ws2, df_usage)

    ws3 = wb.create_sheet("Summary")
    _write_df_to_sheet(ws3, df_summary)

    wb.save(OUTPUT_FILE)
    logger.info("Report saved: %s", OUTPUT_FILE)


# ── Excel formatting helpers (same as integrations script) ──

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2EC"))


def _write_df_to_sheet(ws, df):
    if df.empty:
        ws.append(["No data"])
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for c_idx, _ in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


if __name__ == "__main__":
    main()
