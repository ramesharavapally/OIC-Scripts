#!/usr/bin/env python3
"""
scan_child_integrations.py
==========================
Scans local .iar backup files and reports which child integrations
are invoked from each parent integration.

No API calls. No network access. Reads .iar files only.

How it works:
  Every integration-to-integration invoke inside an OIC Gen3 IAR is
  backed by a .jca file using adapter="collocatedics".  That file
  always contains explicit properties:
      integration_code     - child integration code
      integration_version  - child integration version
      integration_name     - child display name

Output:
  Console summary
  exports/child_integrations_audit.csv
  exports/child_integrations_report.xlsx   (2 sheets: Detail + Summary)

Run:
    python scan_child_integrations.py
"""

import csv
import re
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# ================================================================================
# ▼▼  CONFIGURE THIS  ▼▼

IAR_FOLDER = r"C:\Users\PrasadR\OneDrive - RITE\Desktop\OIC_Scripts\integrations\integration_backups_20260410_185733"

EXTS_DIR    = Path(__file__).parent / "exports"
OUTPUT_CSV  = "child_integrations_audit.csv"
OUTPUT_XLSX = "child_integrations_report.xlsx"

# ▲▲  END OF CONFIGURATION  ▲▲
# ================================================================================


# -- Data model ---------------------------------------------------------------

@dataclass
class ChildRef:
    parent_code:    str
    parent_version: str
    child_code:     str
    child_version:  str
    child_name:     str
    iar_file:       str


# -- IAR scanning -------------------------------------------------------------

_JCA_PROP_RE = re.compile(r'<property\s+name="([^"]+)"\s+value="([^"]+)"', re.I)


def _parent_from_iar(iar_path: Path) -> tuple:
    """
    Derive parent code + version from the IAR filename.
    01_integrations_export.py saves them as  CODE_VV.VV.VVVV.iar
    """
    stem = iar_path.stem
    m = re.search(r"(\d{2}\.\d{2}\.\d{4})$", stem)
    if m:
        ver  = m.group(1)
        code = stem[: -(len(ver) + 1)]
        return code, ver
    return stem, ""


def _scan_iar(iar_path: Path) -> List[ChildRef]:
    pcode, pver = _parent_from_iar(iar_path)
    iar_name    = iar_path.name
    seen        = set()
    refs: List[ChildRef] = []

    try:
        with zipfile.ZipFile(iar_path, "r") as zf:
            for entry in zf.namelist():
                if not entry.lower().endswith(".jca"):
                    continue
                try:
                    content = zf.read(entry).decode("utf-8", errors="replace")
                except Exception:
                    continue

                if "collocatedics" not in content.lower():
                    continue

                props = dict(_JCA_PROP_RE.findall(content))
                ccode = props.get("integration_code", "").strip()
                cver  = props.get("integration_version", "").strip()
                cname = props.get("integration_name", "").strip()

                if not ccode or ccode.upper() == pcode.upper():
                    continue

                key = (ccode.upper(), cver)
                if key in seen:
                    continue
                seen.add(key)

                refs.append(ChildRef(
                    parent_code=pcode,
                    parent_version=pver,
                    child_code=ccode,
                    child_version=cver,
                    child_name=cname,
                    iar_file=iar_name,
                ))

    except zipfile.BadZipFile:
        print(f"  [WARN] Skipping invalid IAR: {iar_name}")

    return refs


# -- Output --------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", fgColor="F2F7FB")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2EC"))


def _write_df_to_sheet(ws, df):
    if df.empty:
        ws.append(["No data"])
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        for c_idx, _ in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if r_idx == 1:
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font   = BODY_FONT
                cell.border = THIN_BORDER
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_csv(refs: List[ChildRef], csv_path: Path):
    fieldnames = [
        "parent_code", "parent_version",
        "child_code", "child_version", "child_name",
        "iar_file",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        for r in refs:
            w.writerow({
                "parent_code":    r.parent_code,
                "parent_version": r.parent_version,
                "child_code":     r.child_code,
                "child_version":  r.child_version,
                "child_name":     r.child_name,
                "iar_file":       r.iar_file,
            })
    print(f"  CSV  -  {csv_path}  ({len(refs)} rows)")


def _write_xlsx(refs: List[ChildRef], xlsx_path: Path):
    columns = ["Parent Code", "Parent Version",
               "Child Code", "Child Version", "Child Name",
               "IAR File"]

    detail_rows = [
        {
            "Parent Code":    r.parent_code,
            "Parent Version": r.parent_version,
            "Child Code":     r.child_code,
            "Child Version":  r.child_version,
            "Child Name":     r.child_name,
            "IAR File":       r.iar_file,
        }
        for r in refs
    ]
    df_detail = (pd.DataFrame(detail_rows, columns=columns)
                 if detail_rows else pd.DataFrame(columns=columns))

    # Summary: child count per parent
    if not df_detail.empty:
        df_summary = (
            df_detail
            .groupby(["Parent Code", "Parent Version"])
            .agg(Child_Count=("Child Code", "count"))
            .reset_index()
            .rename(columns={"Child_Count": "Child Count"})
            .sort_values("Child Count", ascending=False)
            .reset_index(drop=True)
        )
    else:
        df_summary = pd.DataFrame(columns=["Parent Code", "Parent Version", "Child Count"])

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Child Integrations"
    _write_df_to_sheet(ws1, df_detail)

    ws2 = wb.create_sheet("Summary")
    _write_df_to_sheet(ws2, df_summary)

    wb.save(xlsx_path)
    print(f"  XLSX -  {xlsx_path}")


def _print_console(refs: List[ChildRef]):
    if not refs:
        print("\n  No child-integration relationships found.\n")
        return

    groups: dict = {}
    for r in refs:
        key = f"{r.parent_code}|{r.parent_version}"
        groups.setdefault(key, []).append(r)

    print(f"\n{'='*90}")
    print(f"  CHILD INTEGRATION MAP  -  {len(groups)} parent(s), {len(refs)} relationship(s)")
    print("=" * 90)

    for key in sorted(groups):
        parent = groups[key][0]
        print(f"\n  Parent : {parent.parent_code}  v{parent.parent_version}")
        for r in sorted(groups[key], key=lambda x: x.child_code):
            print(f"    +-- {r.child_code:<45}  v{r.child_version}  \"{r.child_name}\"")

    print()


# -- Main ----------------------------------------------------------------------

def main():
    folder = Path(IAR_FOLDER)

    print(f"\n{'='*90}")
    print(f"  OIC CHILD INTEGRATION SCANNER  (IAR only - no API calls)")
    print(f"  Folder : {folder}")
    print(f"  Run at : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*90}\n")

    if not folder.exists():
        print(f"  ERROR: IAR folder not found: {folder}")
        print("         Update IAR_FOLDER at the top of this script.")
        sys.exit(1)

    iar_files = sorted(folder.rglob("*.iar"))
    if not iar_files:
        print(f"  No .iar files found in: {folder}")
        sys.exit(0)

    print(f"  Scanning {len(iar_files)} .iar file(s)...\n")

    all_refs: List[ChildRef] = []
    for iar in iar_files:
        found = _scan_iar(iar)
        if found:
            print(f"  {iar.name}  ->  {len(found)} child integration(s)")
            for r in sorted(found, key=lambda x: x.child_code):
                print(f"      +-- {r.child_code}  v{r.child_version}  \"{r.child_name}\"")
        all_refs.extend(found)

    parent_count = len({(r.parent_code, r.parent_version) for r in all_refs})

    print(f"\n{'-'*90}")
    print(f"  SUMMARY")
    print(f"{'-'*90}")
    print(f"  IAR files scanned         : {len(iar_files)}")
    print(f"  Parents with child calls  : {parent_count}")
    print(f"  Total relationships       : {len(all_refs)}")
    print(f"{'-'*90}")

    _print_console(all_refs)

    EXTS_DIR.mkdir(exist_ok=True)
    _write_csv(all_refs,  EXTS_DIR / OUTPUT_CSV)
    _write_xlsx(all_refs, EXTS_DIR / OUTPUT_XLSX)
    print(f"\n  Done.\n")


if __name__ == "__main__":
    main()
